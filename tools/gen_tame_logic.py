#!/usr/bin/env python3
"""Seed data/tame_logic.json from the 'Ark IDs.xlsx' dependency sheet.

Produces the AP access-logic graph:
  - item recipes: craftable/macro -> requirement expression over other nodes ('+'=AND, '|'=OR)
  - dino tame requirements: dino -> KO/tame method expression
  - ALIAS: every graph node that is a REAL AP-gated engram -> our engrams.json ap_name.
    Nodes with no alias are macros/consumables/resources: they carry no engram of their own and
    just flatten through their recipe requirements (or to nothing = freely available).

Then it FLATTENS each dino's tame requirement to the set of engram ap_names actually needed, and
validates that every referenced engram exists. Run: python tools/gen_tame_logic.py [path-to-xlsx]

This is a SEED tool - data/tame_logic.json is the maintained source afterwards. Re-run only to
re-import from a changed spreadsheet.
"""
import json, os, re, sys

HERE = os.path.dirname(__file__)
ROOT = os.path.normpath(os.path.join(HERE, ".."))
XLSX = sys.argv[1] if len(sys.argv) > 1 else r"C:\Users\justi\Downloads\Ark IDs (1).xlsx"

# sheet token -> our engrams.json ap_name (WITHOUT the "Engram: " prefix). These are the nodes
# that are real AP-gated engrams. CONFIRMED = matches verified against engrams.json name list.
# "??" comment marks a mapping that still needs the user's confirmation (ARK-name ambiguity).
ALIAS = {
    "Campfire": "Campfire", "Cooking Pot": "Cooking Pot", "Bow": "Bow", "Club": "Stone Club",
    "Slingshot": "Slingshot", "Waterskin": "Waterskin", "Refining Forge": "Forge", "Forge": "Forge",
    "Mortar & Pestle": "Mortar And Pestle", "Narcotic": "Narcotic", "Sparkpowder": "Sparkpowder",
    "Gunpowder": "Gunpowder", "Crossbow": "Crossbow", "Fabricator": "Fabricator",
    "Stimulant": "Stimulant", "Bug Repellent": "Bug Repel", "Water Jar": "Water Jar",
    "Tranq Dart": "Tranq Dart", "Electronics": "Electronics", "Polymer": "Polymer",
    "Beer Barrel": "Beer Barrel", "Industrial Forge": "Industrial Forge",
    "Air Conditioner": "Air Conditioner", "Egg Incubator": "Egg Incubator",
    "Stone Arrow": "Arrow Stone", "Tranq Arrow": "Arrow Tranq", "Smithy": "Anvil Bench",
    "Metal Pick": "Metal Pick", "Metal Hatchet": "Metal Hatchet",   # metal tools (key harvest gear)
    "Bola": "Bola", "Preserving Bin": "Preserving Bin",
    "Greenhouse Wall": "Greenhouse Wall", "Greenhouse Ceiling": "Greenhouse Ceiling",
    "Greenhouse Door": "Greenhouse Door", "Greenhouse Doorframe": "Greenhouse Door",
    "Large Crop Plot": "Crop Plot Large", "Medium Crop Plot": "Crop Plot Medium",
    "Basic Rifle Ammo": "Simple Rifle Bullet",     # ?? Longneck ammo == Simple Rifle Bullet
    "Longneck Rifle": "Simple Rifle",              # ?? Longneck == Simple Rifle (vs Machined=assault)
    "Metal Gate Frame": "Metal Gateway",           # ?? gate FRAME == Gateway (Metal Gate = the door)
    "Large Metal Bear Trap": "Bear Trap Large",    # ?? == Bear Trap Large
    "Scuba Mask": "Scuba Helmet Goggles",          # ?? mask == Helmet Goggles
    "Scuba Flippers": "Scuba Boots Flippers",      # ?? == Boots Flippers
    "Scuba Tank": "Scuba Shirt Suit With Tank",    # ?? tank == Shirt Suit With Tank
    "Tree Tap": "Tree Sap Tap",                    # ?? == Tree Sap Tap
    "Electrical Outlet": "Power Outlet",           # ??
    "Electrical Generator": "Power Generator",     # ??
    "Straight Electrical Cable": "Power Cable Straight",   # ??
    "Vertical Electrical Cable": "Power Cable Vertical",   # ??
    # NOTE: Refrigerator / Chemistry Bench are NOT in our engrams.json (excluded) and NO dino tame
    # path needs them (deep tier-7 only), so they carry no engram here. Revisit if the craft-graph
    # scope later needs them.
}
# ---- BOSS / CAVE / TRIBUTE logic (NOT from the tier-gates sheet; best-effort, review) ----
# Extra gear engrams the cave requirements reference (token -> our engram ap_name).
GEAR_ALIAS = {
    "Gas Mask": "Gas Mask", "Ghillie": "Ghillie Shirt", "Fur": "Fur Shirt", "Grenade": "Grenade",
    # "Scuba Tank" alias already in ALIAS above (-> Scuba Shirt Suit With Tank);
    # "Bug Repellent" already aliased in ALIAS above (-> Bug Repel)
}
# their crafting station chain (so "has the engram" also needs the station to make it - avoids a
# softlock where you have the Gas Mask engram but no Fabricator to craft it).
GEAR_RECIPES = {"Gas Mask": "Fabricator", "Ghillie": "Smithy", "Fur": "Smithy",
                "Grenade": "Fabricator"}
# metal tools are crafted from Metal Ingots (Forge) around the smithy/metal-age point - so requiring
# one also requires the Forge, landing them at the right progression tier (not sphere 0).
METAL_TOOL_RECIPES = {"Metal Pick": "Forge", "Metal Hatchet": "Forge"}

# Each Island artifact's CAVE requirement (keyed by the "Artifact: X" short name). BEST-EFFORT
# from general ARK Island knowledge - REVIEW: combat floor + environment gear. Swamp caves need a
# gas mask (user-confirmed); the underwater cave needs scuba (a water mount is implied by it).
# Playtested by Lurch (2026-07-25): most caves are RUN/GRAPPLE/PARACHUTE-through, not combat, so the
# artifact only really needs environmental gear (cold / gas / water), not a KO weapon. Five caves
# need nothing. "Fur Set | (Fur Torso + Otter)" collapses to just Fur: the Fur branch is always
# obtainable, so it's the reachability floor and the Otter branch (a locked tame we intentionally
# keep out of cave logic) never changes what's reachable.
# SUPERSEDED 2026-07-26 (stage 2): cave requirements now come from the SHEET (load_cave_reqs), which
# completes the table Lurch only partly filled in on Discord. His sheet ADDS a Useful<X>Tame mount
# requirement to the five caves he had specified, and gives real requirements for the five he had
# left blank - replacing the placeholder combat floors we were shipping there. Kept here only as the
# fallback if the sheet ever lacks the columns.
CAVE_REQS_FALLBACK = {
    "Brute": "Scuba Tank", "Cunning": "Scuba Tank", "Skylord": "Fur",
    "Strong": "Fur + Grenade", "Immune": "(Gas Mask | Scuba Tank) + Bug Repellent",
    "Hunter": "Crossbow KO", "Massive": "Crossbow KO", "Clever": "Crossbow KO + Scuba Tank",
    "Pack": "Crossbow KO + Fur", "Devourer": "Rifle KO + Gas Mask + Ghillie",
}
# Sheet wording -> our node names, for the cave column specifically.
CAVE_NORMALIZE = {
    "Scuba Set": "Scuba Tank",          # our Scuba Tank alias is the Shirt/Suit-with-tank engram
    "bug repellant": "Bug Repellent",
    "Fur Set": "Fur", "Fur Torso": "Fur",
    "UsefuBrutelTame": "UsefulBruteTame",   # sheet typo
}


def load_cave_reqs(wb):
    """{artifact -> requirement expr} + {Useful<X>Tame -> mount OR-list} from 'tier gates (Base)'."""
    reqs, mounts = {}, {}
    for r in list(wb["tier gates (Base)"].iter_rows(values_only=True))[1:]:
        if len(r) < 14 or not r[10] or str(r[10]).strip() == "Cave":
            continue
        cave = str(r[10]).strip()
        expr = str(r[11] or "").strip()
        for bad, good in CAVE_NORMALIZE.items():
            expr = expr.replace(bad, good)
        reqs[cave] = _norm_macros(expr)
        lst = str(r[13] or "").strip()
        if lst and lst.lower() != "none":
            mounts[f"Useful{cave}Tame"] = _norm_macros(re.sub(r"\s+", " ", lst))
    return reqs, mounts
# artifacts each boss's summon needs (Gamma = artifacts only; goal is any-difficulty).
BOSS_ARTIFACTS = {
    "Broodmother": ["Clever", "Hunter", "Massive"],
    "Megapithecus": ["Pack", "Brute", "Devourer"],
    "Dragon": ["Cunning", "Skylord", "Strong", "Immune"],
}
OVERSEER_BOSSES = ["Broodmother", "Megapithecus", "Dragon"]   # Overseer needs the 3 island bosses

# CAVE-DWELLING tames: taming these means surviving their cave, so their requirement OVERRIDES the
# combat method with a cave floor. Keeps foundational engrams (Mortar, Forge...) from being stranded
# behind a hard cave tame - the fill routes around the dependency.
#
# PASSIVE vs KO matters (review by Lurch9229, 2026-07-23): several of these are PASSIVE tames, so a
# tranq weapon is the wrong gate entirely - you approach unnoticed with Ghillie or Bug Repellent.
# Requiring "Crossbow KO" for them both over-gated (demanded the whole Anvil+Forge+Crossbow chain to
# tame a Dung Beetle) and under-gated (never asked for the gear that is the real barrier).
CAVE_TAMES = {
    # passive - approach gear, NOT a tranq weapon
    "Dung Beetle": "Ghillie | Bug Repellent",
    "Araneo": "Ghillie | Bug Repellent",
    "Onyc": "Ghillie | Bug Repellent",
    "Arthropleura": "Ghillie | Bug Repellent",
    # knockout tames (reviewer raised no objection to these)
    "Pulmonoscorpius": "Crossbow KO",
    "Megalania": "Crossbow KO",
    "Megalosaurus": "Rifle KO",
    # Titanoboa is NOT here on purpose: it's a passive tame that needs a FERTILIZED EGG (breeding,
    # which we don't model) and it's more easily found in the open swamp than in a cave. Listing a
    # combat floor for it was simply wrong. Its tame is instead excluded from carrying progression
    # via NO_TAME_LOGIC in __init__.py, so nothing can be stranded behind an unmodelled breeding
    # requirement (same reasoning as the breed-count milestones).
}

# Explorer notes / dossiers physically inside caves - AUTHORITATIVE from ark.wiki.gg ASE map data
# (Data:Maps/Exploration/The Island/ASE marker groups "dossier cave" / "explorer-note cave"), each
# assigned to its artifact cave. Value = the artifact whose cave it's in -> reuses cave_reqs, so the
# underwater (Cunning) cave notes require scuba automatically. Prevents stranding progression on a
# cave note. Regenerate via the Explorer Map DataMaps API if the wiki updates.
NOTE_CAVES = {
    "Dossier: Allosaurus": "Skylord",
    "Dossier: Araneo": "Clever",
    "Dossier: Carbonemys": "Strong",
    "Dossier: Cnidaria": "Cunning",
    "Dossier: Dilophosaur": "Brute",
    "Dossier: Kaprosuchus": "Hunter",
    "Dossier: Leech": "Hunter",
    "Dossier: Lystrosaurus": "Pack",
    "Dossier: Mammoth": "Pack",
    "Dossier: Manta": "Massive",
    "Dossier: Megaloceros": "Massive",
    "Dossier: Megalodon": "Clever",
    "Dossier: Megalosaurus": "Devourer",
    "Dossier: Mesopithecus": "Devourer",
    "Dossier: Mosasaurus": "Devourer",
    "Dossier: Rex": "Cunning",
    "Dossier: Tapejara": "Immune",
    "Dossier: Terror Bird": "Immune",
    "Dossier: Titanoboa": "Cunning",
    "Dossier: Titanomyrma": "Brute",
    "Dossier: Titanosaur": "Immune",
    "Dossier: Triceratops": "Massive",
    "Dossier: Trilobite": "Skylord",
    "Dossier: Tusoteuthis": "Brute",
    "Helena Note #30": "Devourer",
    "Rockwell Note #29": "Devourer",
    "Mei Yin Note #31": "Devourer",
    "??? Note #1 (idx 508)": "Devourer",
    "Helena Note #1": "Devourer",
    "Helena Note #2": "Hunter",
    "Helena Note #3": "Hunter",
    "Helena Note #4": "Brute",
    "Helena Note #9": "Cunning",
    "Helena Note #11": "Pack",
    "Helena Note #12": "Hunter",
    "Nerva Note #9": "Brute",
    "Mei Yin Note #9": "Cunning",
    "Rockwell Note #18": "Massive",
    "Rockwell Note #17": "Clever",
    "Rockwell Note #6": "Cunning",
    "Rockwell Note #9": "Brute",
}

# tribute organ (inventory-check name prefix, before " xN") -> the roster dino you kill for it.
TRIBUTE_DINO = {
    "Argentavis Talon": "Argentavis", "Sarcosuchus Skin": "Sarcosuchus",
    "Sauropod Vertebra": "Bronto", "Titanoboa Venom": "Titanoboa", "Megalania Toxin": "Megalania",
    "Megalodon Tooth": "Megalodon", "Spinosaurus Sail": "Spino", "Therizino Claws": "Therizinosaurus",
    "Thylacoleo Hook-Claw": "Thylacoleo", "Allosaurus Brain": "Allosaurus",
    "Basilosaurus Blubber": "Basilosaurus", "Giganotosaurus Heart": "Giganotosaurus",
    "Tusoteuthis Tentacle": "Tusoteuthis", "Tyrannosaurus Arm": "Rex", "Yutyrannus Lungs": "Yutyrannus",
}

# nodes that are macros/consumables/resources with NO engram of their own (flatten via recipe).
# Listed so the validator doesn't flag them as "unmapped". Crops/cooked food are free once their
# station requirement (crop plot / campfire) is met - captured by their recipe rows.
FLATTEN_ONLY = {
    "Crossbow KO", "Bow KO", "Rifle KO", "Deep Dive", "Deep Tame", "Deep Caves", "Use Electricity",
    "Cementing Paste", "Gasoline", "Sweet Veggie Cake", "Cooked Meat", "Cooked Prime Meat",
    "Cooked Fish Meat", "Cooked Prime Fish Meat", "Cooked Meat Jerky", "Cooked Prime Meat Jerky",
    "Citronal", "Savoroot", "Longrass", "Rockarrot",
    "Thatch Foundation", "Stone Foundation", "Wood Foundation", "Greenhouse Foundation",
}


# our roster short-name (dinos.json "Tame: X") -> sheet dino name, where they differ. Roster
# dinos NOT here and not spelled identically fall back to DINO_TIER-derived reqs in the apworld.
DINO_ALIAS = {
    "Compsognathus": "Compy", "Triceratops": "Trike", "Stegosaurus": "Stego",
    "Sarcosuchus": "Sarco", "Woolly Rhino": "Wooly Rhino", "Arthropleura": "Arthropluera",
    "Quetzal": "Quetzl", "Yutyrannus": "Yutyranus", "Thylacoleo": "Thylacolio",
    "Procoptodon": "Procoptrodon", "Pulmonoscorpius": "Pulmonoscorpus",
    "Lystrosaurus": "Lystosaurus", "Mesopithecus": "Mesopithicus",
    "Gigantopithecus": "Gigantopithicus", "Carcharodontosaurus": "Carchardontosaurus",
    "Direbear": "Dire bear", "Giant Bee": "Giant Queen Bee", "Therizinosaurus": "Therizino",
}


# ---- Lurch's KILL logic + combat macros ('KillTame Logic' sheet) ---------------------------------
# The sheet writes some macro names inconsistently. Normalising is safe (these are unambiguous
# spelling/spacing/casing variants of a macro that exists); ANY other unknown token fails the build
# rather than being silently dropped, because a dropped requirement makes a check EASIER than
# intended - the exact bug this import is meant to fix.
MACRO_NORMALIZE = {
    "UseArrow": "UseArrows", "Use Arrows": "UseArrows",
    "Use Pistol": "UsePistol", "Use Rifle": "UseRifle", "Use Shotgun": "UseShotgun",
    "useMelee": "UseMelee", "Use Melee": "UseMelee",
    "AdvMelee": "UseAdvMelee", "Use AdvMelee": "UseAdvMelee",
    "Shallow Tames": "ShallowFightTames", "ShallowCombatTames": "ShallowFightTames",
    "Deep Tames": "DeepFightTames", "Deep KO": "DeepKO", "Deep Dive": "DeepDive",
    # item_recipes keys these WITHOUT a space. Written spaced they resolve to nothing at all, and
    # an unresolved token silently becomes "no requirement" - which is how Pulmonoscorpius,
    # Megalania and Megalosaurus ended up with completely free tames.
    "Bow KO": "BowKO", "Crossbow KO": "CrossbowKO", "Rifle KO": "RifleKO",
    "Ride Basilosaurus": "RideBasilosaurus",
    "Metal Pickaxe": "Metal Pick", "Stone Pickaxe": "Stone Pick",
    "UseGrapple": "CanGrapple", "Casteroides": "Castoroides", "Icthysaurus": "Ichthyosaurus",
    "RideAnky": "RideAnkylosaurus", "RideArthroplura": "RideArthropleura",
    "RideBrontosaurus": "RideBronto", "RideCasteroides": "RideCastoroides",
    "RideDunkleosteus": "RideDunkle", "RideIguanadon": "RideIguanodon",
    "RideTapajara": "RideTapejara", "RideIcthysaurus": "RideIchthyosaurus",
    "RideIcthyosaurus": "RideIchthyosaurus", "RideWoollyRhino": "RideWoolly Rhino",
    "RideTherizinosaurus": "RideTherizinosaurus", "RideDireBear": "RideDirebear",
    "RideCarchardontosaurus": "RideCarcharodontosaurus", "RideSarco": "RideSarcosuchus",
    "RideStego": "RideStegosaurus", "RideTrike": "RideTriceratops",
    "RideMosasaur": "RideMosasaur", "RidePulmonoscorpius": "RidePulmonoscorpius",
}
# Rows whose expression is MALFORMED in the sheet (unbalanced parentheses, a stray '|' inside a
# macro name). Corrected explicitly so the fix is visible and reviewable instead of being guessed by
# the parser. Report these upstream so Lurch can clean the sheet.
KILL_FIX = {
    "Pteranodon": "BasicFightTames | UseArrows | UsePistol | UseRifle | UseShotgun "
                  "| (UseMelee + (UseNet | Bola))",
    "Tusotuethis": "DeepDive + (DeepFightTames | RideBasilosaurus)",
    "Achatina": "BasicFightTames | UseArrows | UsePistol | UseRifle | UseShotgun",
}
# Combat-tier macros live in these columns of their sheets: (sheet, {column index: macro name}).
FIGHT_MACRO_COLS = [("KillTame Logic", {8: "ShallowTames", 9: "ShallowFightTames",
                                        10: "DeepTames", 11: "DeepFightTames",
                                        13: "BasicFightTames", 14: "MediumFightTames",
                                        15: "StrongFightTames", 16: "InsaneFightTames"})]
# Titanosaur is KILL-ONLY in our data (it only temp-tames in ASE, so there is no "Tame: Titanosaur"
# item for RideTitanosaur to reference). Drop that one OR-branch; StrongFightTames still offers
# RideRex/Therizinosaurus/Spino/Yutyrannus/Thylacoleo/Rhyniognatha.
RIDE_DROP = {"RideTitanosaur"}

# Weapon/tool engrams the combat macros reference. The sheet uses ARK's display wording; these map
# to our engrams.json names (verified against the name list, not guessed).
COMBAT_ALIAS = {
    "Pike": "Pike", "Spear": "Spear", "Sword": "Sword", "Sickle": "Metal Sickle",
    "Stone Hatchet": "Stone Hatchet", "Club": "Stone Club",
    "Chainsaw": "Chain Saw", "Grappling Hook": "Grappling Hook",
    "Simple Pistol": "Pistol", "Fabricated Pistol": "Machined Pistol",
    "Simple Shotgun": "Simple Shotgun", "Pump Action Shotgun": "Machined Shotgun",
    "Simple Shotgun Ammo": "Simple Shotgun Bullet",
    "Simple Bullet": "Simple Bullet", "Advanced Bullet": "Advanced Bullet",
    "Fabricated Sniper Rifle": "Machined Sniper", "Advanced Sniper Bullet": "Advanced Sniper Bullet",
    "Simple Rifle Ammo": "Simple Rifle Bullet",
    "Harpoon Launcher": "Harpoon Gun", "Net Projectile": "Net Gun Ammo",
    "Cryopod": "Empty Cryopod",
}
# Macro rows on the 'Logic Macros' sheet that we already model better elsewhere, or that name a
# creature rather than an engram. Skipped so they don't shadow the existing definitions.
MACRO_SKIP = {"BowKO", "CrossbowKO", "RifleKO"}   # already exist as "Bow KO"/"Crossbow KO"/"Rifle KO"
# Tokens that carry no AP engram: ARK auto-grants them at level 1 (Stone Pick, Torch - see
# gen_engrams EXCLUDE), or we simply do not model that structure. A branch requiring one is
# always satisfiable, which is correct for the auto-granted tools.
FREE_TOKENS = {"Stone Pick", "Torch", "Wooden Tree Platform", "Metal Tree Platform",
               "Stone Irrigation Pipe - Intake", "Stone Irrigation Pipe - Tap",
               "Metal Irrigation Pipe - Intake", "Metal Irigation Pipe - Tap"}


def load_logic_macros(wb):
    """{macro -> expression} from the 'Logic Macros' sheet (UseArrows, CanFly, DeepDive, ...)."""
    out = {}
    for r in list(wb["Logic Macros"].iter_rows(values_only=True))[1:]:
        if not r or not r[0] or len(r) < 2 or not r[1]:
            continue
        name = str(r[0]).strip()
        if name in MACRO_SKIP:
            continue
        out[name] = _norm_macros(str(r[1]))
    return out


def _norm_macros(expr: str) -> str:
    """Apply the macro spelling table to one expression (longest names first so 'Use Rifle' wins
    over a bare 'Rifle' fragment)."""
    if not expr:
        return ""
    s = str(expr).strip()
    for bad in sorted(MACRO_NORMALIZE, key=len, reverse=True):
        s = re.sub(rf"(?<![A-Za-z]){re.escape(bad)}(?![A-Za-z])", MACRO_NORMALIZE[bad], s)
    for drop in RIDE_DROP:                       # remove the branch AND its dangling separator
        s = re.sub(rf"\s*\|\s*{re.escape(drop)}(?![A-Za-z])", "", s)
        s = re.sub(rf"(?<![A-Za-z]){re.escape(drop)}\s*\|\s*", "", s)
    return s.strip()


def load_kill_logic(wb):
    """{creature -> requirement expression} + {macro -> expression} from the KillTame sheet."""
    ws = wb["KillTame Logic"]
    rows = list(ws.iter_rows(values_only=True))
    kills, macros = {}, {}
    for r in rows[1:]:
        name = r[0] if r else None
        if not name or str(name).strip() == "Kill Logic":
            continue
        name = str(name).strip()
        raw = KILL_FIX.get(name, r[1] if len(r) > 1 else None)
        kills[name] = _norm_macros(_norm(raw))
    for sheet, cols in FIGHT_MACRO_COLS:
        for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
            for ci, macro in cols.items():
                if ci < len(r) and r[ci] and macro not in macros:
                    macros[macro] = _norm_macros(str(r[ci]))
    return kills, macros


def load_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["tier gates (Base)"]
    rows = list(ws.iter_rows(values_only=True))
    items, dinos = {}, {}
    for r in rows[1:]:
        it, ireq, itier = r[0], r[1], r[2]
        dn, dreq, dtier = r[4], r[5], r[6]
        if it:
            items[str(it).strip()] = _norm(ireq)
        if dn:
            dinos[str(dn).strip()] = _norm(dreq)
    return items, dinos, wb


def _norm(expr):
    """Normalize a requirement expression: expand the '/'-style OR shorthands, drop 'None'."""
    if not expr or str(expr).strip().lower() in ("none", ""):
        return ""
    s = str(expr).strip()
    s = s.replace("Medium/Large Crop Plot", "(Medium Crop Plot | Large Crop Plot)")
    s = s.replace("Thatch/Wood/Stone Foundation",
                  "(Thatch Foundation | Wood Foundation | Stone Foundation)")
    return s


def tokens(expr):
    return [t.strip() for t in re.split(r"[+|()]", expr) if t.strip()]


def flatten(node, recipes, seen=None):
    """Return the set of engram ap_names required to obtain/use `node`, recursing recipes.
    A node contributes its own engram (if aliased) AND everything its recipe needs."""
    seen = seen or set()
    if node in seen:                       # cycle guard
        return set()
    seen = seen | {node}
    out = set()
    if node in ALIAS:
        out.add(ALIAS[node])
    for dep in tokens(recipes.get(node, "")):
        out |= flatten(dep, recipes, seen)
    return out


def main():
    items, dinos, wb = load_xlsx(XLSX)
    kills, fight_macros = load_kill_logic(wb)
    fight_macros.update(load_logic_macros(wb))   # UseArrows/CanFly/DeepDive/... macros
    cave_reqs, cave_mounts = load_cave_reqs(wb)   # stage 2: caves from the sheet
    cave_reqs = cave_reqs or dict(CAVE_REQS_FALLBACK)
    fight_macros.update(cave_mounts)              # Useful<Cave>Tame mount lists
    eng = json.load(open(os.path.join(ROOT, "data", "engrams.json"), encoding="utf-8"))
    engset = {e["ap_name"].replace("Engram: ", "") for e in eng["engrams"]}

    # Bow + Tranq Arrow is a valid EARLIER tranq method (crafts in inventory - no Smithy/Forge) for
    # anything a plain Crossbow KO tames. Treat a bare "Crossbow KO" as "Bow KO | Crossbow KO" so the
    # Bow is a real early method AND classifies as progression (never stranded on a late-only check).
    # Only the bare method is relaxed; "Crossbow KO + <gear>" and "Rifle KO" stay as-is (need more).
    def _allow_bow(expr):
        e = (expr or "").strip()
        return "BowKO | CrossbowKO" if e in ("Crossbow KO", "CrossbowKO") else expr
    dinos = {k: _allow_bow(v) for k, v in dinos.items()}
    cave_tames = {k: _allow_bow(_norm_macros(v)) for k, v in CAVE_TAMES.items()}

    # Creatures the sheet leaves blank fall back to DINO_TIER, and a tier-0 creature ends up with NO
    # requirement at all. That is fine for a Dodo, and a real hole for anything OTHER RULES DEPEND
    # ON: cave_reqs for Brute and Cunning both offer "| Diplocaulus" as an alternative to full dive
    # gear, so a free Diplocaulus made every deep-water cave and the twelve notes inside them look
    # sphere-0. A Diplocaulus is a knockout tame - tranq arrows, so narcotics, so a Mortar & Pestle.
    #
    # Only creatures that GATE something else belong here; leaving a trivial tame ungated is correct.
    TAME_FALLBACK = {
        "Diplocaulus": "BowKO | CrossbowKO",         # KO tame; gates cave_reqs Brute + Cunning
    }
    for name, expr in TAME_FALLBACK.items():
        if not (dinos.get(name) or "").strip():
            dinos[name] = expr

    # sanity: every ALIAS target (incl. gear) must be a real engram
    bad_alias = {k: v for k, v in {**ALIAS, **GEAR_ALIAS, **COMBAT_ALIAS}.items()
                 if v not in engset}
    # every referenced token must be aliased, flatten-only, or a recipe node
    all_tokens = set()
    for req in list(items.values()) + list(dinos.values()):
        all_tokens |= set(tokens(req))
    known = set(ALIAS) | FLATTEN_ONLY | set(items)
    unknown = sorted(all_tokens - known)

    # flatten each dino to its engram requirement set
    dino_reqs = {d: sorted(flatten_dino(req, items)) for d, req in dinos.items()}

    print("=== ALIAS targets missing from engrams.json (MUST FIX) ===")
    for k, v in bad_alias.items():
        print(f"  {k!r} -> {v!r}  (not an engram)")
    if not bad_alias:
        print("  (none)")
    print("\n=== tokens referenced but neither aliased nor flatten-only nor a recipe node ===")
    for t in unknown:
        print("  ", t)
    if not unknown:
        print("  (none)")

    # ---- validate the kill logic: every token must resolve to something we can compile ----
    # Known = a macro (fight tiers / Use* / Can*), an aliased engram, a recipe node, a Ride<X> or a
    # bare creature name. Anything else is a typo we must NOT silently drop (it would make the check
    # easier than intended), so it fails the build.
    roster = json.load(open(os.path.join(ROOT, "data", "dinos.json"), encoding="utf-8"))["dinos"]
    def _key(s):
        return re.sub(r"[^a-z]", "", str(s).lower())
    roster_keys = {_key(d.get("name") or (d.get("ap_name") or "").replace("Tame: ", ""))
                   for d in roster}
    macro_names = (set(fight_macros) | set(items) | set(ALIAS) | FLATTEN_ONLY
                   | set(GEAR_ALIAS) | set(COMBAT_ALIAS))
    # Validate only what the KILL/CAVE rules can actually reach. A macro nothing references (the
    # farming chain, say) may mention structures we do not model; flagging those as build errors
    # would block on data we never evaluate.
    def _reachable(seed_exprs):
        seen, stack = set(), list(seed_exprs)
        while stack:
            for t in tokens(stack.pop()):
                if t in seen:
                    continue
                seen.add(t)
                if t in fight_macros:
                    stack.append(fight_macros[t])
                elif t in items:
                    stack.append(items[t])
        return seen
    reachable = _reachable(list(kills.values()) + list(cave_reqs.values()))

    # TAME rules were never validated - only kills and caves were - so a misspelt macro in
    # dino_tame_raw or CAVE_TAMES resolved to "no requirement" and shipped. Three did exactly
    # that. Every token a tame rule uses must be a macro, a recipe, an alias, a roster creature
    # or an explicitly free token; anything else fails the build.
    tame_unknown = {}
    for owner, expr in list(dinos.items()) + list(cave_tames.items()):
        for t in tokens(expr):
            if t in macro_names or t in FREE_TOKENS or t in fight_macros:
                continue
            base = t[4:] if t.startswith("Ride") else t
            if _key(base) in roster_keys or _key(base) in alias_keys:
                continue
            tame_unknown.setdefault(t, []).append(owner)
    if tame_unknown:
        print("\n=== TAME LOGIC: unresolvable tokens (these become NO REQUIREMENT) ===")
        for t, owners in sorted(tame_unknown.items()):
            print(f"  {t!r:24} used by: {', '.join(sorted(owners)[:6])}")
        raise SystemExit("refusing to write tame_logic.json with unresolvable tame tokens")
    alias_keys = {_key(v) for v in DINO_ALIAS.values()} | {_key(k) for k in DINO_ALIAS}
    kill_unknown, unreachable_bad = {}, {}
    for owner, expr in list(kills.items()) + list(fight_macros.items()):
        for t in tokens(expr):
            if t in macro_names or t in FREE_TOKENS:
                continue
            bare = t[4:] if t.startswith("Ride") else t          # Ride<X> -> creature
            if _key(bare) in roster_keys or _key(bare) in alias_keys:
                continue
            (kill_unknown if t in reachable else unreachable_bad).setdefault(t, []).append(owner)

    print("\n=== KILL LOGIC: unresolvable tokens (MUST FIX - add to MACRO_NORMALIZE/KILL_FIX) ===")
    for t, who in sorted(kill_unknown.items()):
        print(f"  {t!r}  used by: {', '.join(who[:4])}")
    if not kill_unknown:
        print("  (none)")
    if unreachable_bad:
        print("\n  (informational) tokens only in macros the kill/cave rules never reach:")
        for t, who in sorted(unreachable_bad.items()):
            print(f"    {t!r}  in {', '.join(sorted(set(who)))}")
    print(f"\nkill rows imported: {len(kills)} | combat macros: {len(fight_macros)}")

    print("\n=== dino tame -> flattened engram requirements ===")
    for d in sorted(dino_reqs):
        print(f"  {d:20} {dino_reqs[d]}")

    out = {"_comment": "Seeded from Ark IDs.xlsx by tools/gen_tame_logic.py. AP tame/craft logic. "
                       "item_recipes + dino_tame_raw + alias are the SOURCE. The apworld compiles "
                       "dino_tame_raw into boolean AP rules (AND='+', OR='|'), expanding macros via "
                       "item_recipes and mapping engram nodes via alias.",
           "item_recipes": {**items, **GEAR_RECIPES, **METAL_TOOL_RECIPES, **fight_macros},
           "dino_tame_raw": dinos, "alias": {**ALIAS, **GEAR_ALIAS, **COMBAT_ALIAS},
           "dino_alias": DINO_ALIAS,
           "_kill_reqs": "Lurch's per-creature KILL requirements ('KillTame Logic' sheet). Same "
                         "grammar as dino_tame_raw. Ride<X> = tame X AND hold X's saddle engram; a "
                         "BARE creature name = just tame it (no saddle needed to fight on it).",
           "kill_reqs": kills,
           "cave_reqs": cave_reqs, "boss_artifacts": BOSS_ARTIFACTS,
           "overseer_bosses": OVERSEER_BOSSES, "tribute_dino": TRIBUTE_DINO,
           "cave_tames": cave_tames, "note_caves": NOTE_CAVES,
           "_dino_tame_engrams_ORasAND": "CONSERVATIVE approximation only (OR flattened to AND, so "
                       "it OVER-requires): use for a quick eyeball / validation, NOT as the rule.",
           "dino_tame_engrams_conservative": dino_reqs}
    dst = os.path.join(ROOT, "data", "tame_logic.json")
    json.dump(out, open(dst, "w", encoding="utf-8"), indent=2)
    print(f"\nwrote {dst}")


def flatten_dino(req, recipes):
    out = set()
    for t in tokens(req):
        out |= flatten(t, recipes)
    return out


if __name__ == "__main__":
    main()
