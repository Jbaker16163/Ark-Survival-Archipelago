"""Append another map's explorer notes to data/locations.json.

Note ids are POSITIONAL - 8740000 + the in-game note index - so unlike engrams and creatures there
is nothing to allocate: the index IS the id. That only works because ARK's note indices are unique
across maps, which was checked rather than assumed (Island 0-1231 and Scorched 90-862 overlap in
exactly zero places).

The reference workbook's Note ID column is trustworthy here, unlike its Item Class column: all 232
indices we harvested for the Island appear in its Island set, so the two agree completely on the
map we can actually verify.

Two types are skipped: "Discovery" (HLN-A) and "Genesis 2 Chronicles". They are Genesis-era notes
that do not exist in ASE, and the Island set excludes them for the same reason.

Names must be unique across the whole AP location table, and each map ships its own Helena/Rockwell
series - "Helena Note #1" exists on both the Island and Scorched Earth. Non-Island notes are
therefore suffixed with the map name. Dossiers are not: a dossier names a creature and is unique.

    python tools/add_map_notes.py --map scorched
    python tools/add_map_notes.py --map scorched --write
"""
import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import checklist_schema as S                                            # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIRS = [os.path.join(ROOT, "data"), os.path.join(ROOT, "apworld", "ark_ase", "data")]
DEFAULT_REF = os.path.join(ROOT, "tools", "reference", "ARK_Creatures_Checklist.xlsx")

# Genesis-era note types that do not exist in ASE on any map.
SKIP_TYPES = {"Discovery", "Genesis 2 Chronicles"}

# reference "Map" cell -> our map key
SHEET_MAP = {"The Island": "island", "Scorched Earth": "scorched", "Aberration": "aberration",
             "Extinction": "extinction", "Genesis: Part 1": "genesis1",
             "Genesis: Part 2": "genesis2", "Ragnarok": "ragnarok", "The Center": "center",
             "Valguero": "valguero", "Crystal Isles": "crystalisles",
             "Lost Island": "lostisland", "Fjordur": "fjordur"}


def load(name):
    with open(os.path.join(DATA_DIRS[0], name), encoding="utf-8") as fh:
        return json.load(fh)


def save(name, obj):
    for d in DATA_DIRS:
        if not os.path.isdir(d):
            continue
        p = os.path.join(d, name)
        with open(p, "w", encoding="utf-8") as fh:
            json.dump(obj, fh, indent=2)
            fh.write("\n")
        print(f"   wrote {p}")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--map", required=True, help="map key to import notes for, e.g. scorched")
    ap.add_argument("--hits", help="ArkAP_note_hits.jsonl from the server. Every note index the "
                                   "game actually handed out is logged there, so this reports any "
                                   "index the reference workbook does not know about - the workbook "
                                   "has gaps (note 1216 appears in no map's list).")
    ap.add_argument("--reference", default=DEFAULT_REF)
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")

    maps_json = load("maps.json")
    known = {m["key"]: m for m in maps_json["maps"]}
    if a.map not in known:
        sys.exit(f"unknown map key '{a.map}'")
    display = known[a.map]["display"]
    sheet_names = [k for k, v in SHEET_MAP.items() if v == a.map]

    locs = load("locations.json")
    doss = locs["location_categories"]["dossiers"]
    entries = doss["entries"]
    have_idx = {e["note_index"] for e in entries}
    # names must be unique across EVERY location category, not just the notes
    have_names = {e["name"] for blk in locs["location_categories"].values()
                  for e in blk.get("entries", [])}

    wb = load_workbook(a.reference, read_only=True, data_only=True)
    rows = list(wb["Explorer Notes"].iter_rows(values_only=True))
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    im, iid, ity, ino = (hdr.index("Map"), hdr.index("Note ID"), hdr.index("Type"),
                         hdr.index("Note"))

    lo, hi = S.ID_BLOCKS["note"]
    added, skipped_type, dupes = [], 0, []
    for r in rows[1:]:
        # A few notes are shared and the sheet writes them as "Ragnarok / Crystal Isles / Fjordur",
        # so match any component rather than the whole cell.
        if not r or r[im] is None:
            continue
        cell_maps = [p.strip() for p in str(r[im]).split("/")]
        if not any(m in sheet_names for m in cell_maps):
            continue
        try:
            idx = int(r[iid])
        except (TypeError, ValueError):
            continue
        kind = str(r[ity] or "").strip()
        if kind in SKIP_TYPES:
            skipped_type += 1
            continue
        if idx in have_idx:
            dupes.append(idx)                      # already ours (shared note, or re-run)
            continue
        note = str(r[ino] or "").strip()
        if kind == "Dossier":
            name = "Dossier: " + re.sub(r"\s*Dossier$", "", note)
        else:
            name = f"{note} ({display})" if a.map != "island" else note
        if name in have_names:
            sys.exit(f"name collision even after qualifying: {name!r} (note index {idx})")
        nid = lo + idx
        if not (lo <= nid <= hi):
            sys.exit(f"note index {idx} falls outside the id block {lo}-{hi}")
        have_names.add(name)
        have_idx.add(idx)
        added.append({"id": nid, "name": name, "note_index": idx})

    if a.hits:
        seen = set()
        for line in open(a.hits, encoding="utf-8"):
            line = line.strip()
            if not line:
                continue
            try:
                seen.add(int(json.loads(line)["note_index"]))
            except (ValueError, KeyError):
                continue
        have = {e["note_index"] for e in entries} | {e["note_index"] for e in added}
        gaps = sorted(seen - have)
        print(f"note hits      : {len(seen)} distinct index(es) seen in game")
        print(f"UNKNOWN to us  : {len(gaps)} -> {gaps}")
        if gaps:
            print("   These fired in game but exist in no map's list. They are real, collectible")
            print("   notes we would never award a check for. Add them by hand once identified.")

    print(f"map            : {a.map} ({display})")
    print(f"skipped types  : {skipped_type} ({', '.join(sorted(SKIP_TYPES))} - not in ASE)")
    print(f"already ours   : {len(dupes)}")
    print(f"TO ADD         : {len(added)}")
    for e in added[:6]:
        print(f"     {e['id']}  idx {e['note_index']:5}  {e['name']}")
    if len(added) > 6:
        print(f"     ... and {len(added) - 6} more")

    if not a.write:
        print("\nDRY RUN. Re-run with --write to apply.")
        return
    if not added:
        print("\nnothing to add")
        return

    entries.extend(added)
    # Sorted by id keeps the Island notes in their existing relative order, which is what the
    # dossier_checks window walks after the map filter has been applied.
    entries.sort(key=lambda e: e["id"])
    doss["map"] = "multiple"                       # was "TheIsland"; it is no longer one map's list
    save("locations.json", locs)

    content = maps_json["content"]
    b = content.setdefault(a.map, {"items": [], "locations": []})
    for e in added:
        b["locations"].append(e["id"])
    for bucket in content.values():
        bucket["items"] = sorted(set(bucket["items"]))
        bucket["locations"] = sorted(set(bucket["locations"]))
    save("maps.json", maps_json)
    print(f"\n{len(added)} note(s) added; dossiers now {len(entries)}")


if __name__ == "__main__":
    main()
