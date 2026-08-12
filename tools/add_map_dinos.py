"""Append another map's creatures to data/dinos.json from a /dumpdinos harvest.

The plugin's taming and kill gates key on the runtime **DinoNameTag** - what DinoNameTagField()
returns - and that string is NOT derivable from anything we already have. Checked against our 105
verified creatures, "entity id minus _Character_BP_C" is wrong for 18 of them: Bronto is Sauropod,
Titanoboa is BoaFrill, Araneo is SpiderS, Procoptodon is Kangaroo. So the tag has to be harvested
from a running server, exactly like engram classes had to be dumped rather than reconstructed.

Harvest it by running the map and using the in-game command:

    /dumpdinos            (5 km radius around you; run it from a few spots)
    cheat DestroyWildDinos    (passive Die harvest catches everything loaded at once)

That appends to ArkAP_dino_classes.jsonl next to the plugin. Feed that file here.

Each source is used only for what it is authoritative about:
  * harvest  - the DinoNameTag and the real spawn class
  * reference workbook - display name, tameable/rideable, and which maps the creature is on
  * engrams.json - the saddle's engram_class, matched to the creature's entity prefix

    python tools/add_map_dinos.py <ArkAP_dino_classes.jsonl> --maps scorched,ragnarok
    ... same again with --write
"""
import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import checklist_schema as S                                            # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIRS = [os.path.join(ROOT, "data"), os.path.join(ROOT, "apworld", "ark_ase", "data")]
DEFAULT_REF = os.path.join(ROOT, "tools", "reference", "ARK_Creatures_Checklist.xlsx")


def load(name):
    with open(os.path.join(DATA_DIRS[0], name), encoding="utf-8") as fh:
        return json.load(fh)


def save(name, obj):
    for d in DATA_DIRS:
        if not os.path.isdir(d):
            continue
        p = os.path.join(d, name)
        with open(p, "w", encoding="utf-8") as fh:
            json.dump(obj, fh, indent=2)
            fh.write("\n")
        print(f"   wrote {p}")


def ticked(v):
    t = str(v or "").strip()
    return bool(t) and t.upper() not in ("0", "-", "N", "NO", "FALSE", "X", "NONE")


def entity_prefix(entity):
    """'SpineyLizard_Character_BP_C' -> 'SpineyLizard'. Used only to find the saddle engram."""
    return re.sub(r"_Character.*$", "", str(entity or "").strip())


def next_free(used, base, top):
    """Allocate ABOVE everything in use, never into a gap.

    The gaps are not spare space: gen_dinos.py assigns ids as BASE + roster index, so the holes are
    slots it will hand to some other creature the next time it runs. Filling one guarantees a
    collision later - two creatures with the same tame_loc, which AP sees as one location and the
    plugin fires for the wrong species. Appending past the maximum can never collide."""
    i = max(used) + 1 if used else base
    if i < base:
        i = base
    if i > top:
        sys.exit(f"id block {base}-{top} exhausted")
    return i



def _resolve(classes, tag, by_entity, by_prefix):
    """Find the reference row for a harvested creature, loosest match last.

    The harvest and the sheet do not always spell the class the same way, and both spellings are
    legitimate - the game ships variant classes the sheet only lists once:
      * Jugbug_Oil_Character_BP_C   vs the sheet's Jugbug_Character_BaseBP_C
      * camelsaurus_Character_BP_C  vs the sheet's Camelsaurus_Character_BP_C  (game lowercases it)
    Returns (row, how) so the caller can print which rule fired - a loose match is worth eyeballing.
    """
    for c in classes:                                    # 1. exact class
        if c in by_entity:
            return by_entity[c], "class"
    lower = {k.lower(): v for k, v in by_entity.items()}
    for c in classes:                                    # 2. class, case-insensitive
        if c.lower() in lower:
            return lower[c.lower()], "class/case"
    for c in classes:                                    # 3. entity prefix, then trim variants
        parts = entity_prefix(c).lower().split("_")
        for n in range(len(parts), 0, -1):
            hit = by_prefix.get("_".join(parts[:n]))
            if hit is not None:
                return hit, "prefix"
    hit = by_prefix.get(tag.lower())                     # 4. the tag itself
    if hit is not None:
        return hit, "tag"
    return None, ""


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("harvest", nargs="+",
                    help="one or more ArkAP_dino_classes.jsonl files from /dumpdinos. Pass every "
                         "map's harvest together - which maps a creature belongs to comes from the "
                         "reference workbook, never from which file it was seen in, so merging "
                         "harvests cannot mis-tag anything.")
    ap.add_argument("--maps", required=True, help="map keys these creatures are on")
    ap.add_argument("--reference", default=DEFAULT_REF)
    ap.add_argument("--write", action="store_true")
    a = ap.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")

    want = S.parse_map_list(a.maps)
    maps_json = load("maps.json")
    known = {m["key"] for m in maps_json["maps"]}
    for k in want:
        if k not in known:
            sys.exit(f"unknown map key '{k}' - not on the Maps sheet")

    # ---- harvest: one record per canon tag (Wyvern ships Fire/Lightning/Poison classes, all one
    # creature and all one tag, so collapse them) --------------------------------------------
    harvested = {}
    for path in a.harvest:
        n0 = len(harvested)
        with open(path, encoding="utf-8") as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    r = json.loads(line)
                except ValueError:
                    continue
                tag = (r.get("canon") or r.get("tag") or "").strip()
                if tag:
                    harvested.setdefault(tag, []).append(r.get("class", ""))
        print(f"harvest         : {os.path.basename(path)} -> {len(harvested) - n0} new tag(s)")
    print(f"                  {len(harvested)} distinct tag(s) across "
          f"{sum(len(v) for v in harvested.values())} class record(s)")

    dinos_json = load("dinos.json")
    dinos = dinos_json["dinos"]
    have_tags = {d.get("dino_tag") for d in dinos}

    # ---- reference: entity class -> row -----------------------------------------------------
    wb = load_workbook(a.reference, read_only=True, data_only=True)
    rows = list(wb["Creatures"].iter_rows(values_only=True))
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    ie, ia = hdr.index("Entity ID"), hdr.index("In ASE Steam")
    itm, ird = hdr.index("Tamable"), hdr.index("Ridable")
    by_entity, by_prefix = {}, {}
    for r in rows[1:]:
        if r and r[0] and str(r[ie] or "").strip():
            ent = str(r[ie]).strip()
            by_entity[ent] = r
            by_prefix.setdefault(entity_prefix(ent).lower(), r)


    # ---- saddle engram lookup: entity prefix -> engram_class --------------------------------
    saddles = {}
    for e in load("engrams.json")["engrams"]:
        m = re.search(r"PrimalItemArmor_(.+?)Saddle", e["engram_class"])
        if m:
            saddles[m.group(1).lower()] = e["engram_class"]

    used_items = {d["id"] for d in dinos if d.get("id")}
    used_tame = {d["tame_loc"] for d in dinos if d.get("tame_loc")}
    used_kill = {d["kill_loc"] for d in dinos if d.get("kill_loc")}

    by_tag = {d.get("dino_tag"): d for d in dinos}
    added, no_ref, not_ase, already = [], [], [], 0
    widen = []                       # creatures we already have, but on a map not yet recorded
    for tag, classes in sorted(harvested.items()):
        if tag in have_tags:
            already += 1
            # Do NOT just skip. Harvesting Scorched then Ragnarok would otherwise leave Wyvern
            # tagged scorched only, because the second run sees it as already ours - so a Ragnarok
            # slot would lose a creature standing in front of the player.
            row, _how = _resolve(classes, tag, by_entity, by_prefix)
            if row is not None and ticked(row[ia]):
                present = [key for col, key in S.REFERENCE_MAP_COLUMNS.items()
                           if col in hdr and key in known and ticked(row[hdr.index(col)])]
                mine = [k for k in want if k in present]
                d = by_tag.get(tag) or {}
                ids = [d[f] for f in ("id", "tame_loc", "kill_loc") if d.get(f)]
                gap = [k for k in mine
                       if any(i not in set(maps_json["content"].get(k, {}).get("items", []))
                              and i not in set(maps_json["content"].get(k, {}).get("locations", []))
                              for i in ids)]
                if gap and ids:
                    widen.append({"name": str(row[0]).strip(), "tag": tag, "maps": mine, "ids": ids})
            continue
        row, how = _resolve(classes, tag, by_entity, by_prefix)
        if row is None:
            no_ref.append((tag, classes[0] if classes else ""))
            continue
        if not ticked(row[ia]):
            not_ase.append(tag)
            continue
        name = str(row[0]).strip()
        present = [key for col, key in S.REFERENCE_MAP_COLUMNS.items()
                   if col in hdr and key in known and ticked(row[hdr.index(col)])]
        # Membership is the INTERSECTION, never the whole --maps list. Harvesting on Scorched can
        # turn up a creature that only lives on one of the maps you named (Karkinos is Aberration
        # and Ragnarok, not Scorched), and tagging it with both would put it in a Scorched pool it
        # can never be tamed in - an unreachable location that fails the seed.
        mine = [k for k in want if k in present]
        if not mine:
            continue
        tameable = ticked(row[itm])
        entry = {}
        kill = next_free(used_kill, S.ID_BLOCKS["kill_loc"][0], S.ID_BLOCKS["kill_loc"][1])
        used_kill.add(kill)
        if tameable:
            iid = next_free(used_items, S.ID_BLOCKS["dino_item"][0], S.ID_BLOCKS["dino_item"][1])
            tame = next_free(used_tame, S.ID_BLOCKS["tame_loc"][0], S.ID_BLOCKS["tame_loc"][1])
            used_items.add(iid)
            used_tame.add(tame)
            entry = {"id": iid, "ap_name": "Tame: " + name, "dino_tag": tag,
                     "tame_loc": tame, "kill_loc": kill}
            # Always carry the key, null when there is no saddle - that is what import_checklist
            # writes, and a mismatch would make the workbook round trip rewrite the file forever.
            sc = saddles.get(entity_prefix(row[ie]).lower()) if ticked(row[ird]) else None
            entry["saddle_class"] = sc
            if ticked(row[ird]) and not sc:
                entry["_no_saddle"] = True          # rideable but no saddle engram found
        else:
            entry = {"name": name, "dino_tag": tag, "kill_loc": kill, "tameable": False}
        added.append({"entry": entry, "maps": mine, "name": name, "how": how,
                      "rideable": ticked(row[ird])})

    print(f"already ours    : {already}"
          + (f"  ({len(widen)} of them gain a map)" if widen else ""))
    for w in widen:
        print(f"     WIDEN {w['name']:18} tag={w['tag']:16} -> {', '.join(w['maps'])}")
    print(f"TO ADD          : {len(added)}  -> maps: {', '.join(want)}")
    for x in added:
        e = x["entry"]
        kind = "tame+kill" if "id" in e else "kill only"
        maps_txt = ",".join(x["maps"])
        sad = ("saddle" if e.get("saddle_class") else
               "NO SADDLE FOUND" if e.pop("_no_saddle", False) else "-")
        loose = "" if x["how"] == "class" else f"  [matched by {x['how']}]"
        print(f"     {x['name']:20} tag={e['dino_tag']:18} {kind:10} {sad:16} "
              f"maps={maps_txt}{loose}")
    if no_ref:
        print(f"\nharvested but not in the reference ({len(no_ref)}) - skipped:")
        for t, c in no_ref[:10]:
            print(f"     tag={t:20} class={c}")
    if not_ase:
        print(f"\nnot in ASE per the reference ({len(not_ase)}): {not_ase}")

    missing_saddle = [x["name"] for x in added
                      if x["rideable"] and not x["entry"].get("saddle_class")]
    if missing_saddle:
        print(f"\nRIDEABLE WITH NO SADDLE ENGRAM ({len(missing_saddle)}): {missing_saddle}")
        print("   Correct for Wyvern/Griffin/Phoenix (no saddle exists). For anything else, add "
              "the saddle engram first with add_map_engrams.py, or Ride<X> logic can never resolve.")
    print("\nSTILL TO AUTHOR BY HAND after this: a tame requirement for each new creature in "
          "tame_logic.json, and a tier. Without a rule the creature is sphere-0 (tameable with "
          "nothing), which is the same silent-permissive bug the Diplocaulus token had.")

    if not a.write:
        print("\nDRY RUN. Re-run with --write to apply.")
        return
    if not added:
        print("\nnothing to add")
        return

    for x in added:
        x["entry"].pop("_no_saddle", None)
        dinos.append(x["entry"])
    for w in widen:                              # existing creature, newly-known map
        for k in w["maps"]:
            b = maps_json["content"].setdefault(k, {"items": [], "locations": []})
            for i in w["ids"]:
                b["items" if 8732001 <= i <= 8732999 else "locations"].append(i)
    base = dinos_json.get("_comment", "").split(" Appended:")[0]
    dinos_json["_comment"] = (
        f"{base} Appended: {len(added)} creature(s) for {', '.join(want)} by add_map_dinos.py "
        f"(ids allocated ABOVE the generated block - gen_dinos.py's ids are positional, so filling "
        f"a gap would collide the next time it runs).")
    save("dinos.json", dinos_json)

    content = maps_json["content"]
    for x in added:
        e = x["entry"]
        for k in x["maps"]:                     # per-creature, not the whole --maps list
            b = content.setdefault(k, {"items": [], "locations": []})
            if e.get("id"):
                b["items"].append(e["id"])
            for loc in ("tame_loc", "kill_loc"):
                if e.get(loc):
                    b["locations"].append(e[loc])
    for b in content.values():
        b["items"] = sorted(set(b["items"]))
        b["locations"] = sorted(set(b["locations"]))
    save("maps.json", maps_json)
    print(f"\n{len(added)} creature(s) added; dinos.json now has {len(dinos)}")


if __name__ == "__main__":
    main()
