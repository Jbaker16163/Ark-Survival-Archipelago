"""Fill in which maps each creature / alpha / engram actually exists on.

Membership used to be one map per id, which is wrong: 104 of the Island's 113 ASE creatures also
spawn on Ragnarok and every Island engram is craftable there. Tagging a Rex "island" would delete
it from a Ragnarok slot's pool even though it is standing right in front of the player.

The presence matrix comes from the reference workbook (tools/reference/ARK_Creatures_Checklist.xlsx
by default) - a 14-map grid for Creatures, Alpha Predators, Engrams and Artifacts. This tool reads
that grid, matches it against our own data files by name, and widens data/maps.json accordingly.

It is ADDITIVE and DRY BY DEFAULT:
  * nothing is written without --write
  * existing membership is never removed, only widened, so a bad name match can lose no content
  * ids it cannot resolve are reported and left exactly as they are
  * categories it does not own (notes, bosses, crates, inventory, levels, milestones) are untouched

    python tools/backfill_map_membership.py            # report only
    python tools/backfill_map_membership.py --write     # apply to data/maps.json
"""
import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import checklist_schema as S                                            # noqa: E402

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# The data files live in TWO mirrors and both are loaded at runtime - data/ by the tools and the
# plugin, apworld/ark_ase/data/ by the apworld. Writing only one leaves the generator reading a
# stale copy, which looks exactly like the backfill silently doing nothing.
DATA_DIRS = [os.path.join(ROOT, "data"), os.path.join(ROOT, "apworld", "ark_ase", "data")]
DATA = DATA_DIRS[0]
DEFAULT_REF = os.path.join(ROOT, "tools", "reference", "ARK_Creatures_Checklist.xlsx")

# Our short names vs the reference sheet's. The reference spells Arthropleura "Arthropluera" - kept
# verbatim on purpose, because the point of this table is to match THAT file, not to be correct.
# "Pachy" is Pachycephalosaurus; Pachyrhinosaurus is a different animal and must not collide.
CREATURE_ALIASES = {
    "Angler": "Anglerfish", "Arthropleura": "Arthropluera", "Bronto": "Brontosaurus",
    "Carno": "Carnotaurus", "Compsognathus": "Compy", "Dunkle": "Dunkleosteus",
    "Jellyfish": "Cnidaria", "Mosasaur": "Mosasaurus", "Pachycephalosaurus": "Pachy",
    "Sarcosuchus": "Sarco", "Spino": "Spinosaurus",
}


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())


def ticked(v):
    """A presence cell. The sheet uses check marks, but tolerate the usual spellings so a
    hand-edited column still reads correctly."""
    t = str(v or "").strip()
    return bool(t) and t.upper() not in ("0", "-", "N", "NO", "FALSE", "X", "NONE")


def load(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as fh:
        return json.load(fh)


def sheet(wb, title):
    """(header list, list of row tuples) for a sheet, or (None, []) if it is absent."""
    if title not in wb.sheetnames:
        return None, []
    rows = list(wb[title].iter_rows(values_only=True))
    if not rows:
        return None, []
    return [str(h) if h is not None else "" for h in rows[0]], rows[1:]


def presence(hdr, row, known_keys, ase_only=True):
    """Map keys this row is present on. Returns None when the row is not in ASE at all - such a
    creature must never enter an ASE pool, whatever its map columns say."""
    if ase_only and "In ASE Steam" in hdr and not ticked(row[hdr.index("In ASE Steam")]):
        return None
    out = set()
    for col, key in S.REFERENCE_MAP_COLUMNS.items():
        if col in hdr and key in known_keys and ticked(row[hdr.index(col)]):
            out.add(key)
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--reference", default=DEFAULT_REF, help="reference workbook (14-map matrix)")
    ap.add_argument("--write", action="store_true", help="apply to data/maps.json (default: report)")
    a = ap.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl is required: pip install openpyxl")
    if not os.path.isfile(a.reference):
        sys.exit(f"reference workbook not found: {a.reference}\n"
                 f"pass --reference, or copy it to {DEFAULT_REF}")

    wb = load_workbook(a.reference, read_only=True, data_only=True)
    maps_json = load("maps.json")
    known = {m["key"] for m in maps_json["maps"]} - {"any"}
    content = maps_json["content"]

    # id -> maps we are about to add. Kept separate from `content` so the report can show exactly
    # what would change before anything is written.
    add = {}
    unresolved, not_in_ase = [], []

    def widen(ids, keys, kind):
        for i in ids:
            if i is not None:
                add.setdefault((kind, int(i)), set()).update(keys)

    # ---- creatures: the dino unlock item, plus its tame and kill checks ---------------------
    hdr, rows = sheet(wb, "Creatures")
    ref = {}
    if hdr:
        for r in rows:
            if r and r[0]:
                ref[norm(r[0])] = r
    matched = 0
    for e in load("dinos.json")["dinos"]:
        name = e.get("name") or e["ap_name"].replace("Tame: ", "")
        row = ref.get(norm(CREATURE_ALIASES.get(name, name))) or ref.get(norm(name))
        if row is None:
            unresolved.append(f"creature {name}")
            continue
        keys = presence(hdr, row, known)
        if keys is None:
            not_in_ase.append(f"creature {name}")
            continue
        matched += 1
        widen([e.get("id")], keys, "items")
        widen([e.get("tame_loc"), e.get("kill_loc")], keys, "locations")
    print(f"creatures      : {matched} matched, {len(unresolved)} unresolved")

    # ---- alpha kills -------------------------------------------------------------------------
    hdr, rows = sheet(wb, "Alpha Predators")
    aref = {norm(r[0]): r for r in rows if r and r[0]} if hdr else {}
    alphas = load("locations.json")["location_categories"].get("alpha_kills", {}).get("entries", [])
    n = 0
    for e in alphas:
        short = e["name"].replace("Killed: ", "").replace("Alpha ", "")
        row = next((r for k, r in aref.items() if norm(short) in k or k in norm(e["name"])), None)
        if row is None:
            unresolved.append(f"alpha {e['name']}")
            continue
        keys = presence(hdr, row, known)
        if keys is None:
            not_in_ase.append(f"alpha {e['name']}")
            continue
        n += 1
        widen([e["id"]], keys, "locations")
    print(f"alpha kills    : {n} matched of {len(alphas)}")

    # ---- engrams ------------------------------------------------------------------------------
    # Engrams already sit under "any", which is strictly wider than any per-map list, so they are
    # only reported here - widening them would NARROW them. Left alone deliberately.
    hdr, rows = sheet(wb, "Engrams")
    if hdr:
        isl = sum(1 for r in rows if r and r[0] and (presence(hdr, r, known) or set()) >= {"island"})
        rag = sum(1 for r in rows if r and r[0] and (presence(hdr, r, known) or set()) >= {"ragnarok"})
        print(f'engrams        : reference says island={isl} ragnarok={rag}; ours are tagged "any" '
              f"(already wider than any map list) - not touched")

    # ---- report --------------------------------------------------------------------------------
    widened = new_ids = 0
    per_map = {}
    for (kind, i), keys in add.items():
        cur = {k for k, b in content.items() if i in b.get(kind, [])} - {"any"}
        if "any" in {k for k, b in content.items() if i in b.get(kind, [])}:
            continue                                    # already every map; nothing to add
        gained = keys - cur
        if gained:
            widened += 1
            for k in gained:
                per_map[k] = per_map.get(k, 0) + 1
        if not cur:
            new_ids += 1
    print(f"\n{widened} ids would gain membership ({new_ids} had none at all)")
    for k, v in sorted(per_map.items(), key=lambda kv: -kv[1]):
        print(f"   +{v:4} ids -> {k}")
    if unresolved:
        print(f"\nUNRESOLVED ({len(unresolved)}) - membership left untouched:")
        for u in unresolved:
            print("   ", u)
    if not_in_ase:
        print(f"\nnot in ASE per the reference ({len(not_in_ase)}) - left untouched:")
        for u in not_in_ase:
            print("   ", u)

    if not a.write:
        print("\nDRY RUN. Re-run with --write to apply to data/maps.json")
        return
    for (kind, i), keys in add.items():
        for k in keys:
            b = content.setdefault(k, {"items": [], "locations": []})
            if i not in b[kind]:
                b[kind].append(i)
    for b in content.values():
        b["items"] = sorted(set(b["items"]))
        b["locations"] = sorted(set(b["locations"]))
    for d in DATA_DIRS:
        if not os.path.isdir(d):
            continue
        dst = os.path.join(d, "maps.json")
        with open(dst, "w", encoding="utf-8") as fh:
            json.dump(maps_json, fh, indent=2)
            fh.write("\n")
        print(f"\nwrote {dst}")
    for k in sorted(content):
        print(f"   {k:14} items={len(content[k]['items']):4} locations={len(content[k]['locations']):4}")


if __name__ == "__main__":
    main()
